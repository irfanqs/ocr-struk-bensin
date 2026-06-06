from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


COLUMNS = [
    "source_pdf",
    "source_folder",
    "page",
    "tanggal",
    "jam",
    "produk",
    "volume_liter",
    "harga_per_liter",
    "total_rupiah",
    "spbu",
    "no_nota",
    "rfid",
    "nopol",
    "nopol_raw_text",
    "operator",
    "selected_rotation",
    "needs_review",
    "review_reason",
    "tanggal_confidence",
    "produk_confidence",
    "volume_liter_confidence",
    "total_rupiah_confidence",
    "tanggal_raw_text",
    "produk_raw_text",
    "volume_liter_raw_text",
    "total_rupiah_raw_text",
    "full_text",
    "cache_json",
]


def export_rows(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "HASIL OCR"
    ws.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in rows:
        ws.append([row.get(column) for column in COLUMNS])
        if row.get("needs_review"):
            for cell in ws[ws.max_row]:
                cell.fill = review_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for index, column in enumerate(COLUMNS, start=1):
        width = max(len(column) + 2, 12)
        if column in {"source_pdf", "source_folder", "full_text", "cache_json", "review_reason"}:
            width = 42
        ws.column_dimensions[get_column_letter(index)].width = width

    wb.save(output_path)
