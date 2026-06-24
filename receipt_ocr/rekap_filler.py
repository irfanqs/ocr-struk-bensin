from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


MONTHS = {
    1: "JANUARI", 2: "FEBRUARI", 3: "MARET", 4: "APRIL",
    5: "MEI", 6: "JUNI", 7: "JULI", 8: "AGUSTUS",
    9: "SEPTEMBER", 10: "OKTOBER", 11: "NOVEMBER", 12: "DESEMBER",
}
MONTH_ALIASES = {
    "JAN": 1, "JANUARI": 1, "FEB": 2, "FEBRUARI": 2,
    "MAR": 3, "MARET": 3, "APR": 4, "APRIL": 4,
    "MEI": 5, "JUN": 6, "JUNI": 6, "JUL": 7, "JULI": 7,
    "AGU": 8, "AGUSTUS": 8, "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OKT": 10, "OKTO": 10, "OKTOBER": 10,
    "NOV": 11, "NOPEMBER": 11, "NOVEMBER": 11,
    "DES": 12, "DESEMBER": 12,
}
FUEL_ALIASES = {
    "pertalite": "pertalite", "pertalita": "pertalite",
    "pertamax": "pertamax",
    "solar": "solar", "bio solar": "solar", "biosolar": "solar",
    "dexlite": "dexlite",
    "pertamina dex": "pertamina dex", "dex": "pertamina dex",
}

_STATUS_LABELS = {
    "tanggal_tidak_ditemukan": "Tanggal tidak terbaca",
    "nopol_tidak_ditemukan": "No Plat tidak terbaca",
    "jenis_bbm_tidak_dikenali": "Jenis BBM tidak dikenali",
    "kolom_nopol_bbm_tidak_ditemukan": "Kolom plat/BBM tidak ada di sheet",
    "baris_tanggal_tidak_ditemukan": "Baris tanggal tidak ada di sheet",
    "baris_tanggal_penuh": "Semua baris tanggal sudah terisi",
}

PLATE_PATTERN = re.compile(r"\b(?:L\s*\d{4}\s*[A-Z]{1,3}|AE\s*8719\s*SQ)\b")


def fill_rekap(rows: list[dict[str, Any]], template_path: Path, output_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(template_path)
    logs: list[dict[str, Any]] = []
    column_cache: dict[str, list[ColumnPair]] = {}

    for row in rows:
        log: dict[str, Any] = {
            "source_pdf": row.get("source_pdf"),
            "page": row.get("page"),
            "tanggal": row.get("tanggal"),
            "nopol": row.get("nopol"),
            "produk": row.get("produk"),
            "volume_liter": row.get("volume_liter"),
            "total_rupiah": row.get("total_rupiah"),
            "status": None, "reason": None, "sheet": None,
            "excel_row": None, "liter_col": None, "rupiah_col": None,
        }

        trx_date = parse_date(row.get("tanggal"))
        trx_date = apply_source_period(trx_date, row.get("source_pdf"))
        if trx_date:
            log["tanggal"] = trx_date.isoformat()
        plate = normalize_plate(row.get("nopol"))
        log["nopol"] = plate
        fuel = normalize_fuel(row.get("produk"))
        volume = to_number(row.get("volume_liter"))
        amount = to_number(row.get("total_rupiah"))

        if not trx_date:
            logs.append(_fail(log, "tanggal_tidak_ditemukan"))
            continue
        if not plate:
            logs.append(_fail(log, "nopol_tidak_ditemukan"))
            continue
        if not fuel:
            logs.append(_fail(log, "jenis_bbm_tidak_dikenali"))
            continue

        sheet_name = f"{MONTHS[trx_date.month]} {trx_date.year}"
        if sheet_name not in wb.sheetnames:
            logs.append(_fail(log, f"sheet_tidak_ada:{sheet_name}"))
            continue

        ws = wb[sheet_name]
        if sheet_name not in column_cache:
            column_cache[sheet_name] = _build_column_pairs(ws)

        pair = _find_column_pair(column_cache[sheet_name], plate, fuel)
        if not pair:
            logs.append(_fail(log, "kolom_nopol_bbm_tidak_ditemukan"))
            continue

        row_range = _find_date_rows(ws, trx_date)
        if not row_range:
            logs.append(_fail(log, "baris_tanggal_tidak_ditemukan"))
            continue

        target_row, fill_status = _find_target_row(ws, row_range, pair, volume, amount)

        if fill_status == "already_filled":
            log.update({
                "status": "already_filled",
                "reason": f"sudah terisi: {ws.cell(target_row, pair.liter_col).value} L / Rp {ws.cell(target_row, pair.amount_col).value}",
                "sheet": sheet_name, "excel_row": target_row,
                "liter_col": pair.liter_col, "rupiah_col": pair.amount_col,
            })
            logs.append(log)
            continue

        if target_row is None:
            logs.append(_fail(log, "baris_tanggal_penuh"))
            continue

        ws.cell(target_row, pair.liter_col).value = volume
        ws.cell(target_row, pair.amount_col).value = amount
        _copy_style_from_above(ws, target_row, pair.liter_col)
        _copy_style_from_above(ws, target_row, pair.amount_col)

        is_partial = volume is None or amount is None
        log.update({
            "status": "partial_fill" if is_partial else "filled",
            "reason": "volume atau jumlah tidak terbaca dari OCR, isi manual" if is_partial else None,
            "sheet": sheet_name, "excel_row": target_row,
            "liter_col": pair.liter_col, "rupiah_col": pair.amount_col,
        })
        logs.append(log)

    _write_raw_sheet(wb, rows)
    _write_log_sheet(wb, logs)
    _write_review_sheet(wb, logs)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return logs


class ColumnPair:
    def __init__(self, plates: set[str], fuel: str | None, liter_col: int, amount_col: int) -> None:
        self.plates = plates
        self.fuel = fuel
        self.liter_col = liter_col
        self.amount_col = amount_col


def _build_column_pairs(ws: Any) -> list[ColumnPair]:
    pairs: list[ColumnPair] = []
    col = 2
    while col < ws.max_column:
        current = _cell_value(ws, 6, col)
        next_value = _cell_value(ws, 6, col + 1)
        if not current or _normalize_text(next_value) != "rp":
            col += 1
            continue
        current_norm = _normalize_text(current)
        header_current = _cell_value(ws, 5, col)
        header_next = _cell_value(ws, 5, col + 1)
        fuel = normalize_fuel(header_current if current_norm.startswith("liter") else current)
        plates = _extract_plates(" ".join(str(v or "") for v in (header_current, header_next)))
        if fuel and plates:
            pairs.append(ColumnPair(plates, fuel, col, col + 1))
        col += 2
    return pairs


def _find_column_pair(pairs: list[ColumnPair], plate: str, fuel: str) -> ColumnPair | None:
    exact = [p for p in pairs if plate in p.plates and p.fuel == fuel]
    if exact:
        return exact[0]
    for p in pairs:
        if plate in p.plates and p.fuel and fuel in {p.fuel, normalize_fuel(p.fuel)}:
            return p
    return None


def _find_date_rows(ws: Any, target_date: date) -> tuple[int, int] | None:
    for row in range(1, ws.max_row + 1):
        value = _cell_value(ws, row, 1)
        if parse_date(value) != target_date:
            continue
        for merged_range in ws.merged_cells.ranges:
            if ws.cell(row, 1).coordinate in merged_range:
                return merged_range.min_row, merged_range.max_row
        return row, row
    return None


def _find_target_row(
    ws: Any, row_range: tuple[int, int], pair: ColumnPair, volume: Any, amount: Any
) -> tuple[int | None, str]:
    empty_row = None
    for row in range(row_range[0], row_range[1] + 1):
        liter_val = ws.cell(row, pair.liter_col).value
        amount_val = ws.cell(row, pair.amount_col).value
        if liter_val in (None, "") and amount_val in (None, ""):
            if empty_row is None:
                empty_row = row
        elif volume is not None and amount is not None:
            try:
                if abs(float(liter_val or 0) - float(volume)) < 0.01 and abs(float(amount_val or 0) - float(amount)) < 1:
                    return row, "already_filled"
            except (TypeError, ValueError):
                pass
    return (empty_row, "empty") if empty_row is not None else (None, "full")


def _cell_value(ws: Any, row: int, col: int) -> Any:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return cell.value
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None


def _copy_style_from_above(ws: Any, row: int, col: int) -> None:
    if row <= 1:
        return
    source = ws.cell(row - 1, col)
    target = ws.cell(row, col)
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format


def _write_log_sheet(wb: Any, logs: list[dict[str, Any]]) -> None:
    if "OCR LOG" in wb.sheetnames:
        del wb["OCR LOG"]
    ws = wb.create_sheet("OCR LOG")
    columns = ["status", "reason", "sheet", "excel_row", "liter_col", "rupiah_col",
               "tanggal", "nopol", "produk", "volume_liter", "total_rupiah", "source_pdf", "page"]
    ws.append(columns)
    for log in logs:
        ws.append([log.get(c) for c in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_raw_sheet(wb: Any, rows: list[dict[str, Any]]) -> None:
    if "DATA OCR" in wb.sheetnames:
        del wb["DATA OCR"]
    ws = wb.create_sheet("DATA OCR")
    columns = [
        "source_pdf", "source_folder", "page", "tanggal", "jam", "produk",
        "volume_liter", "harga_per_liter", "total_rupiah", "spbu", "no_nota",
        "rfid", "nopol", "nopol_raw_text", "operator", "selected_rotation",
        "needs_review", "review_reason", "tanggal_confidence", "produk_confidence",
        "volume_liter_confidence", "total_rupiah_confidence", "tanggal_raw_text",
        "produk_raw_text", "volume_liter_raw_text", "total_rupiah_raw_text",
        "full_text", "cache_json",
    ]
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c) for c in columns])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


_REVIEW_FILL = PatternFill("solid", fgColor="FFD7D7")
_ALREADY_FILL = PatternFill("solid", fgColor="D7F5D7")
_PARTIAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


def _write_review_sheet(wb: Any, logs: list[dict[str, Any]]) -> None:
    flagged = [l for l in logs if l.get("status") in ("skipped", "already_filled", "partial_fill")]
    if not flagged:
        return
    if "PERLU REVIEW" in wb.sheetnames:
        del wb["PERLU REVIEW"]
    ws = wb.create_sheet("PERLU REVIEW")
    columns = ["status", "keterangan", "tanggal", "nopol", "produk",
               "volume_liter", "total_rupiah", "sheet", "excel_row", "source_pdf", "page"]
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
    for log in flagged:
        status = log.get("status")
        reason = _STATUS_LABELS.get(log.get("reason", ""), log.get("reason", ""))
        ws.append([status, reason] + [log.get(c) for c in columns[2:]])
        row_fill = _ALREADY_FILL if status == "already_filled" else (_PARTIAL_FILL if status == "partial_fill" else _REVIEW_FILL)
        for cell in ws[ws.max_row]:
            cell.fill = row_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 28 if col in ("keterangan", "source_pdf") else 16


def _fail(log: dict[str, Any], reason: str) -> dict[str, Any]:
    log["status"] = "skipped"
    log["reason"] = reason
    return log


def _extract_plates(text: str) -> set[str]:
    plates = set()
    for match in re.finditer(PLATE_PATTERN, text.upper()):
        plate = normalize_plate(match.group(0))
        if plate:
            plates.add(plate)
    return plates


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def normalize_plate(value: Any) -> str | None:
    if value is None:
        return None
    match = PLATE_PATTERN.search(str(value).upper())
    if not match:
        return None
    return re.sub(r"[^A-Z0-9]", "", match.group(0))


def normalize_fuel(value: Any) -> str | None:
    if value is None:
        return None
    text = _normalize_text(value)
    for alias, normalized in FUEL_ALIASES.items():
        if alias in text:
            return normalized
    return None


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt).date()
        except ValueError:
            pass
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def apply_source_period(value: date | None, source_pdf: Any) -> date | None:
    source_period = _infer_period_from_source(source_pdf)
    if not value or not source_period:
        return value
    month, year = source_period
    try:
        return date(year, month, value.day)
    except ValueError:
        return value


def _infer_period_from_source(source_pdf: Any) -> tuple[int, int] | None:
    if source_pdf is None:
        return None
    text = _normalize_text(source_pdf).upper()
    year_match = re.search(r"\b(20\d{2})\b", text)
    if not year_match:
        return None
    year = int(year_match.group(1))
    for name, month in MONTH_ALIASES.items():
        if re.search(rf"\b{name}\b", text):
            return month, year
    return None


def to_number(value: Any) -> float | int | None:
    if isinstance(value, int | float):
        return value
    if value is None:
        return None
    text = re.sub(r"[^0-9,.-]", "", str(value).strip())
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number
