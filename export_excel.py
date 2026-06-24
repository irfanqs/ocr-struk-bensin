#!/usr/bin/env python3
"""
Program 2: Baca JSON cache hasil OCR → export ke Excel rekap.

Cara pakai:
  python export_excel.py                          # baca output/cache/, buat hasil_ocr_struk.xlsx
  python export_excel.py --fill-rekap             # sekaligus isi workbook rekap bulanan
  python export_excel.py --cache-dir output/cache --excel output/hasil.xlsx
  python export_excel.py --fill-rekap \\
    --rekap-template "REKAP PENGGUNAAN_BBM_PERIODE_BULAN_NOVEMBER 2025-2.xlsx" \\
    --rekap-output output/rekap_terisi.xlsx
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from receipt_ocr.rekap_filler import (
    _STATUS_LABELS,
    fill_rekap,
    normalize_plate,
)


COLUMNS = [
    "source_pdf", "source_folder", "page",
    "tanggal", "jam", "produk", "volume_liter", "harga_per_liter", "total_rupiah",
    "spbu", "no_nota", "rfid", "nopol", "nopol_raw_text", "operator",
    "selected_rotation", "needs_review", "review_reason",
    "tanggal_confidence", "produk_confidence", "volume_liter_confidence", "total_rupiah_confidence",
    "tanggal_raw_text", "produk_raw_text", "volume_liter_raw_text", "total_rupiah_raw_text",
    "full_text", "cache_json",
]

REQUIRED_FIELDS = ("tanggal", "produk", "volume_liter", "total_rupiah")


def main() -> None:
    args = build_parser().parse_args()
    cache_dir = Path(args.cache_dir)

    if not cache_dir.exists():
        print(f"ERROR: Folder cache tidak ditemukan: {cache_dir}")
        print("Jalankan ocr.py terlebih dahulu untuk menghasilkan file JSON cache.")
        raise SystemExit(1)

    json_files = sorted(cache_dir.glob("*.json"))
    if not json_files:
        print(f"Tidak ada file JSON di {cache_dir}. Jalankan ocr.py dulu.")
        raise SystemExit(1)

    print(f"Membaca {len(json_files)} file JSON dari {cache_dir} ...")
    rows = [load_json_row(f) for f in json_files]

    excel_path = Path(args.excel)
    export_rows(rows, excel_path)
    print(f"Excel disimpan: {excel_path}  ({len(rows)} baris)")

    if args.fill_rekap:
        template = Path(args.rekap_template)
        output = Path(args.rekap_output)
        if not template.exists():
            print(f"ERROR: Template rekap tidak ditemukan: {template}")
            raise SystemExit(1)
        logs = fill_rekap(rows, template, output)
        print_rekap_summary(logs, str(output))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Export JSON cache hasil OCR ke Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--cache-dir", default="output/cache",
                        help="Folder berisi JSON cache hasil OCR. (default: output/cache)")
    parser.add_argument("--excel", default="output/hasil_ocr_struk.xlsx",
                        help="Path output Excel hasil OCR. (default: output/hasil_ocr_struk.xlsx)")
    parser.add_argument("--fill-rekap", action="store_true",
                        help="Sekaligus isi workbook rekap bulanan dari hasil OCR.")
    parser.add_argument(
        "--rekap-template",
        default="REKAP PENGGUNAAN_BBM_PERIODE_BULAN_NOVEMBER 2025-2.xlsx",
        help="Workbook rekap template yang akan disalin dan diisi.",
    )
    parser.add_argument("--rekap-output", default="output/rekap_terisi.xlsx",
                        help="Path workbook rekap hasil. (default: output/rekap_terisi.xlsx)")
    return parser


def load_json_row(json_path: Path) -> dict[str, Any]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    meta = data.get("_meta", {})
    row: dict[str, Any] = {
        "source_pdf": meta.get("source_pdf"),
        "source_folder": meta.get("source_folder"),
        "page": meta.get("page"),
        "cache_json": str(json_path),
        "selected_rotation": data.get("selected_rotation"),
        "needs_review": bool(data.get("needs_review", True)),
        "review_reason": data.get("review_reason"),
        "full_text": data.get("full_text"),
    }

    for field in ("spbu", "tanggal", "jam", "no_nota", "produk", "harga_per_liter",
                  "volume_liter", "total_rupiah", "rfid", "nopol", "operator"):
        item = data.get(field) if isinstance(data.get(field), dict) else {}
        row[field] = item.get("value")
        row[f"{field}_confidence"] = item.get("confidence")
        row[f"{field}_raw_text"] = item.get("raw_text") or item.get("value")
        row[f"{field}_notes"] = item.get("notes")

    row["nopol"] = normalize_plate(row.get("nopol"))

    if not row["needs_review"]:
        row["needs_review"] = any(_is_low_confidence(data.get(f)) for f in REQUIRED_FIELDS)
        if row["needs_review"] and not row["review_reason"]:
            row["review_reason"] = "field penting kosong atau confidence rendah"

    return row


def _is_low_confidence(item: Any) -> bool:
    if not isinstance(item, dict):
        return True
    if item.get("value") in (None, ""):
        return True
    try:
        return float(item.get("confidence", 0)) < 0.75
    except (TypeError, ValueError):
        return True


def export_rows(rows: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "HASIL OCR"
    ws.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in rows:
        ws.append([row.get(c) for c in COLUMNS])
        if row.get("needs_review"):
            for cell in ws[ws.max_row]:
                cell.fill = review_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(COLUMNS, start=1):
        width = 42 if col in {"source_pdf", "source_folder", "full_text", "cache_json", "review_reason"} else max(len(col) + 2, 12)
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(output_path)


def print_rekap_summary(logs: list[dict[str, Any]], output_path: str) -> None:
    counts: dict[str, int] = {}
    for log in logs:
        s = log.get("status", "unknown")
        counts[s] = counts.get(s, 0) + 1

    print(f"\nRekap disimpan ke: {output_path}")
    print("=" * 50)
    print(f"  Terisi              : {counts.get('filled', 0)}")
    print(f"  Terisi sebagian     : {counts.get('partial_fill', 0)}  ← lihat sheet PERLU REVIEW")
    print(f"  Sudah ada isinya    : {counts.get('already_filled', 0)}  ← lihat sheet PERLU REVIEW")
    print(f"  Dilewati (OCR miss) : {counts.get('skipped', 0)}  ← lihat sheet PERLU REVIEW")
    print("=" * 50)

    skipped = [l for l in logs if l.get("status") == "skipped"]
    if skipped:
        print(f"\nDilewati ({len(skipped)} transaksi):")
        for l in skipped:
            label = _STATUS_LABELS.get(l.get("reason", ""), l.get("reason", ""))
            print(f"  {l.get('tanggal') or '?'} | {l.get('nopol') or '?'} | {label}")

    partial = [l for l in logs if l.get("status") == "partial_fill"]
    if partial:
        print(f"\nTerisi sebagian ({len(partial)} transaksi) — isi manual volume/jumlah di Excel:")
        for l in partial:
            print(f"  Baris {l.get('excel_row')} sheet {l.get('sheet')} | {l.get('nopol')} | {l.get('produk')}")


if __name__ == "__main__":
    main()
